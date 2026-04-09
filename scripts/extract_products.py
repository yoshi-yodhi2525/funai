"""
基礎データ0155.xlsx から製品マスタを抽出して
Zoho CRM 商品モジュールインポート用CSVを生成するスクリプト

使い方:
  pip install openpyxl
  python extract_products.py

出力:
  products_import.csv  → Zoho CRM「商品」モジュールにインポート
  products_debug.txt   → 全シートの生データ（デバッグ用）

Excel 構造 (Sheet3 機器表):
  各カテゴリが横方向に 6列単位で並ぶ
  +0: 品名, +1: 型式, +2: 小売価格(参照), +4: 原価(参照), +5: 歩単(労務率)
"""

import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です: pip install openpyxl")
    sys.exit(1)

BASE_DIR   = Path(__file__).parent.parent
INPUT_FILE = BASE_DIR / "基礎データ0155.xlsx"
OUT_CSV    = Path(__file__).parent / "products_import.csv"
OUT_DEBUG  = Path(__file__).parent / "products_debug.txt"

# Zoho CRM 商品モジュールのフィールド名 (API名)
CSV_FIELDS = [
    "Product_Name",      # 商品名
    "Product_Code",      # 型番
    "Product_Category",  # カテゴリ（シート名/分類）
    "Unit_Price",        # 定価（小売価格）
    "Cost",              # 原価
    "Description",       # 備考（歩単等）
]


def load_workbook():
    if not INPUT_FILE.exists():
        print(f"ファイルが見つかりません: {INPUT_FILE}")
        sys.exit(1)
    print(f"読み込み中: {INPUT_FILE}")
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    print(f"シート: {wb.sheetnames}")
    return wb


def safe_num(val):
    """セル値を数値に変換。失敗時は None"""
    if val is None:
        return None
    try:
        n = float(str(val).replace(",", "").strip())
        return int(n) if n == int(n) else n
    except (ValueError, TypeError):
        return None


def safe_str(val):
    """セル値を文字列に変換"""
    if val is None:
        return ""
    return str(val).strip()


# ============================================================
# Sheet3: 機器表 （横方向に 6列ずつカテゴリが並ぶ）
# ============================================================
def extract_sheet3(ws, products, debug_lines):
    """
    機器表 列レイアウト（K列=11から 8列単位で繰り返し）:
      col+0: 品番 (Product_Code)
      col+1: 品名 (カテゴリ兼用)
      col+2: 型式 (model spec)
      col+3: 単位
      col+4: 小売価格 (Unit_Price)
      col+5: 原価 (Cost)
      col+6: 歩単 (労務比率)
    データは Row3 以降
    """
    debug_lines.append("\n=== Sheet3 (機器表) ===")

    max_col = ws.max_column
    max_row = ws.max_row

    start_col = 11  # K列 (1始まり)
    col_span  = 8   # 次セクションまで8列

    section_count = 0
    product_count = 0

    for base_col in range(start_col, max_col, col_span):
        # Row3の品番セルが空なら終了
        if ws.cell(row=3, column=base_col).value is None:
            break

        section_count += 1
        debug_lines.append(f"\n  セクション開始列: {base_col}")

        for row in range(3, max_row + 1):
            jancode  = safe_str(ws.cell(row=row, column=base_col).value)
            name     = safe_str(ws.cell(row=row, column=base_col + 1).value)
            model    = safe_str(ws.cell(row=row, column=base_col + 2).value)
            unit_str = safe_str(ws.cell(row=row, column=base_col + 3).value)
            price    = safe_num(ws.cell(row=row, column=base_col + 4).value)
            cost     = safe_num(ws.cell(row=row, column=base_col + 5).value)
            labor    = safe_str(ws.cell(row=row, column=base_col + 6).value)

            if not name or name.startswith("─") or name.startswith("-"):
                continue

            full_name = f"{name} {model}".strip() if model else name

            product = {
                "Product_Name":     full_name,
                "Product_Code":     jancode,
                "Product_Category": name,
                "Unit_Price":       price if price is not None else "",
                "Cost":             cost  if cost  is not None else "",
                "Description":      f"型式:{model} 単位:{unit_str} 歩単:{labor}",
            }
            products.append(product)
            product_count += 1
            debug_lines.append(f"    {jancode} | {full_name} | {price}")

    debug_lines.append(f"\n  → {section_count} セクション, {product_count} 件")
    return product_count


# ============================================================
# Sheet4: 電材表 （同様のレイアウト）
# ============================================================
def extract_sheet4(ws, products, debug_lines):
    debug_lines.append("\n=== Sheet4 (電材表) ===")
    max_col = ws.max_column
    max_row = ws.max_row
    start_col = 8   # H列
    col_span  = 6
    count = 0

    for base_col in range(start_col, max_col, col_span):
        category = ""
        for r in range(1, 6):
            val = ws.cell(row=r, column=base_col).value
            if val and str(val).strip():
                category = safe_str(val)
                break
        if not category:
            continue

        for row in range(6, max_row + 1):
            name  = safe_str(ws.cell(row=row, column=base_col).value)
            code  = safe_str(ws.cell(row=row, column=base_col + 1).value)
            price = safe_num(ws.cell(row=row, column=base_col + 2).value)
            cost  = safe_num(ws.cell(row=row, column=base_col + 4).value)

            if not name or name.startswith("─"):
                continue

            products.append({
                "Product_Name":     name,
                "Product_Code":     code,
                "Product_Category": f"電材/{category}",
                "Unit_Price":       price if price is not None else "",
                "Cost":             cost  if cost  is not None else "",
                "Description":      "",
            })
            count += 1

    debug_lines.append(f"  → {count} 件")
    return count


# ============================================================
# Sheet5: 管材表
# ============================================================
def extract_sheet5(ws, products, debug_lines):
    debug_lines.append("\n=== Sheet5 (管材表) ===")
    max_col = ws.max_column
    max_row = ws.max_row
    start_col = 10   # J列
    col_span  = 6
    count = 0

    for base_col in range(start_col, max_col, col_span):
        category = ""
        for r in range(1, 6):
            val = ws.cell(row=r, column=base_col).value
            if val and str(val).strip():
                category = safe_str(val)
                break
        if not category:
            continue

        for row in range(6, max_row + 1):
            name  = safe_str(ws.cell(row=row, column=base_col).value)
            code  = safe_str(ws.cell(row=row, column=base_col + 1).value)
            price = safe_num(ws.cell(row=row, column=base_col + 2).value)
            cost  = safe_num(ws.cell(row=row, column=base_col + 4).value)

            if not name or name.startswith("─"):
                continue

            products.append({
                "Product_Name":     name,
                "Product_Code":     code,
                "Product_Category": f"管材/{category}",
                "Unit_Price":       price if price is not None else "",
                "Cost":             cost  if cost  is not None else "",
                "Description":      "",
            })
            count += 1

    debug_lines.append(f"  → {count} 件")
    return count


# ============================================================
# メイン処理
# ============================================================
def main():
    wb          = load_workbook()
    products    = []
    debug_lines = []

    sheet_extractors = {
        3: extract_sheet3,   # 機器表
        4: extract_sheet4,   # 電材表
        5: extract_sheet5,   # 管材表
    }

    for sheet_idx, extractor in sheet_extractors.items():
        ws_name = wb.sheetnames[sheet_idx - 1] if sheet_idx <= len(wb.sheetnames) else None
        if not ws_name:
            print(f"Sheet{sheet_idx} が存在しません。スキップ。")
            continue
        try:
            ws = wb[ws_name]
            extractor(ws, products, debug_lines)
        except Exception as e:
            print(f"Sheet{sheet_idx} 抽出エラー: {e}")
            debug_lines.append(f"\n[ERROR] Sheet{sheet_idx}: {e}")

    # 重複排除（Product_Code + Product_Name）
    seen    = set()
    unique  = []
    for p in products:
        key = (p["Product_Name"], p["Product_Code"])
        if key not in seen and p["Product_Name"]:
            seen.add(key)
            unique.append(p)

    # CSV 出力
    with open(OUT_CSV, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_FIELDS)
        writer.writeheader()
        writer.writerows(unique)

    # デバッグ出力
    with open(OUT_DEBUG, "w", encoding="utf-8") as f:
        f.write("\n".join(debug_lines))

    print(f"\n[完了] 抽出完了: {len(unique)} 件")
    print(f"   CSV:   {OUT_CSV}")
    print(f"   Debug: {OUT_DEBUG}")
    print()
    print("[Zoho CRM インポート手順]")
    print("  1. Zoho CRM → 商品 モジュールを開く")
    print("  2. 右上「...」→「インポート」")
    print("  3. products_import.csv を選択")
    print("  4. フィールドマッピングを確認してインポート実行")
    print()
    print("[カスタムフィールド追加（見積書モジュール）]")
    print("  Zoho CRM 設定 → モジュールとフィールド → 見積書 に追加が必要なフィールド:")
    print("  ・見積連番      (field_seq_no)    - 1行テキスト")
    print("  ・改訂番号      (field_revision)  - 整数")
    print("  ・見積明細JSON  (field_items_json)- 複数行テキスト（大）")


if __name__ == "__main__":
    main()
